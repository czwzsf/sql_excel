from django import forms
from django.http import JsonResponse
from django.shortcuts import render, HttpResponse, redirect
from openpyxl.reader.excel import load_workbook
from django.views.decorators.csrf import csrf_exempt
from read_excel import models
from django.db.models import Q
import json


# Create your views here.

class sale_data_form(forms.ModelForm):
    class Meta:
        model = models.sale_data
        fields = "__all__"


class parts_data_form(forms.ModelForm):
    class Meta:
        model = models.parts_data
        fields = "__all__"


def upload_excel_sale(request):
    """基于Excel的文件上传"""
    # 1. 获取用户上传的销售数据文件对象
    form = sale_data_form(data=request.POST, files=request.FILES)
    file_object = request.FILES.get("filename", None)
    if not file_object:
        return HttpResponse("没有文件可供上传")
    # 2. 将文件对象传递给openpyxl，由openpyxl来进行读取
    wb = load_workbook(file_object)
    # 3. 将读取出来的第一个表格传入到sheet里面
    sheet = wb.worksheets[0]
    dict_list = []
    for row in sheet.iter_rows(min_row=2):
        for item in range(len(row)):
            text = str(row[item].value)
            dict_list.append(text)
        models.sale_data.objects.create(production_code=dict_list[1], chassis_code=dict_list[2],
                                        offline_data=dict_list[3],
                                        sale_data=dict_list[4], engine_number=dict_list[5], engine_type=dict_list[6],
                                        responsible_organization=dict_list[7], dealer_code=dict_list[8],
                                        dealer_name=dict_list[9], representative_office_code=dict_list[10],
                                        representative_office_name=dict_list[11], cluth=dict_list[12],
                                        gear_box=dict_list[13],
                                        Rear_axle_speed_ratio=dict_list[14], cage=dict_list[15],
                                        spread_of_axles=dict_list[16],
                                        aak_data=dict_list[17])
        dict_list.clear()
        # if form.is_valid():
        #     form.save()
    return HttpResponse("文件上传完成")


def info(request):
    if request.method == "GET":
        return render(request, 'html/read_excel/read_excel.html')
    return render(request, 'html/read_excel/read_excel.html')


def upload_excel_claim(request):
    """基于Excel的索赔数据上传"""
    return None


def upload_excel_parts(request):
    """基于Excel的文件上传"""
    # 1. 获取用户上传的销售数据文件对象
    form = parts_data_form(data=request.POST, files=request.FILES)
    file_object_parts = request.FILES.get("filename", None)
    if not file_object_parts:
        return HttpResponse("没有文件可供上传")
    # 2. 将文件对象传递给openpyxl，由openpyxl来进行读取
    wb = load_workbook(file_object_parts)
    # 3. 将读取出来的第一个表格传入到sheet里面
    sheet = wb.worksheets[0]
    dict_list_parts = []
    for row in sheet.iter_rows(min_row=2):
        for item in range(len(row)):
            text = str(row[item].value)
            dict_list_parts.append(text)
        models.parts_data.objects.create(statistical_year=dict_list_parts[1], parts=dict_list_parts[2],
                                         mis_3=dict_list_parts[3], mis_6=dict_list_parts[4], mis_9=dict_list_parts[5],
                                         mis_12=dict_list_parts[6],
                                         type=dict_list_parts[7], engine_platform=dict_list_parts[8],
                                         let=dict_list_parts[9], purpose=dict_list_parts[10])
        dict_list_parts.clear()
    return HttpResponse("文件上传完成")


@csrf_exempt
def mis_info(request):
    # 用户第一次请求时为
    if request.method == 'GET':
        queryset1 = models.lbjmc.objects.all().order_by('id')
        form = parts_data_form()
        context = {
            "queryset1": queryset1,
            "form": form,
        }
        return render(request, 'html/mis/mis.html', context=context)
    context = {
        "parts": request.POST.get('parts'),
        "platform": request.POST.get('platform'),
        "type": request.POST.get('type'),
        "let": request.POST.get('let'),
        "purpose": request.POST.get('purpose'),
    }
    # 零部件名称
    query = Q(parts=context['parts'])
    # 查询发动机平台
    if context['platform']:
        query = query & Q(engine_platform=context['platform'])
    # 查询车型
    if context['type']:
        query = query & Q(type=context['type'])
    # 查询排放
    if context['let']:
        query = query & Q(let=context['let'])
    # 查询用途
    if context['purpose']:
        query = query & Q(purpose=context['purpose'])
    # 查询前端用户选择
    query_set = models.parts_data.objects.filter(query)
    if query_set:
        queryset1 = models.lbjmc.objects.all().order_by('id')
        # 标签
        legend = ['3MIS', '6MIS', '9MIS', '12MIS']
        # 横坐标
        x_axis = []
        # mis信息
        data_3mis = []
        data_6mis = []
        data_9mis = []
        data_12mis = []
        title = query_set[0].parts
        for form in query_set:
            x_axis.append(form.statistical_year)
            data_3mis.append(form.mis_3)
            data_6mis.append(form.mis_6)
            data_9mis.append(form.mis_9)
            data_12mis.append(form.mis_12)
        series_list = [
            {
                "name": "3MIS",
                "type": "line",
                "data": data_3mis,
                "stack": "Total",
            },
            {
                "name": "6MIS",
                "type": "line",
                "data": data_6mis,
                "stack": "Total",
            },
            {
                "name": "9MIS",
                "type": "line",
                "data": data_9mis,
                "stack": "Total",
            },
            {
                "name": "12MIS",
                "type": "line",
                "data": data_12mis,
                "stack": "Total",
            },
        ]
        context = {
            "parts": request.POST.get('parts'),
            "platform": request.POST.get('platform'),
            "type": request.POST.get('type'),
            "let": request.POST.get('let'),
            "purpose": request.POST.get('purpose'),
            "status": True,
            "title": title,
            "legend": legend,
            "series_list": series_list,
            "x_axis": x_axis,
            "queryset1": queryset1,
        }
        return render(request, 'html/mis/mis_chart.html', context)
    return render(request, 'html/mis/mis_chart.html', context)


def mis_chart(request):
    str1 = "测试js"
    return render(request, 'html/test.html', {'str1': json.dumps(str1)})
